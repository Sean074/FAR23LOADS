"""The methods & limitations statement reaches every export channel (Step G8.3).

Modelled on ``test_ultimate_contract.py``, and for the same reason: a deliverable
that leaves this tool must carry its own basis *in band*. A span-load CSV
forwarded to a stress engineer, or a BDF handed to sbeam, has to say by itself
that its numbers are ULTIMATE, under what category, and what the tool does not
do. An on-page caption does not travel with a downloaded file.

Two failure modes are pinned:

1. **A channel loses the stamp** — the statement is built in one place
   (``report.methods``) but has to be *passed* to each writer, and a writer that
   silently drops its ``header_comment`` argument would go unnoticed.
2. **The stamp breaks the payload** — ``#`` comment lines above a CSV header row
   are only harmless if every reader skips them. A stamped CSV must parse to the
   *same rows* as an unstamped one.
"""

import csv as _csv
import io as _io
import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import sloads.modules  # noqa: F401
from sloads import io
from sloads.export import sbeam_bridge as sb
from sloads.modules.flight_envelope import build_envelope
from sloads.modules.net_loads import build_net_loads
from sloads.registry import run_all_modules
from sloads.report.methods import (
    APPROVED_CORRECTIONS,
    STANDING_DISCLAIMER,
    _standing_limitations,
    bdf_comment_block,
    csv_comment_block,
    methods_statement,
    strip_comment_lines,
)

_EXAMPLES = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "examples")
_GA = os.path.join(_EXAMPLES, "ga6_normal.project.json")
_CONCEPT = os.path.join(_EXAMPLES, "concept_regional_jet.project.json")


def _project(path):
    return io.load_project(path)


def _wing_net(path):
    p = _project(path)
    if p.envelope is None:
        p.envelope = build_envelope(p)
    return p, build_net_loads(p).wing_net


# --------------------------------------------------------------------------- #
# The statement itself
# --------------------------------------------------------------------------- #
def test_statement_carries_every_required_block():
    """The content blocks of plan §5 / SUMMARY_REPORT.md §4.6, each identified by
    its heading — including the standing disclaimer (``STATUS:``, item 9)."""
    text = methods_statement(_project(_GA), tool_version="0.3.0", scope="full case set")
    for heading in ("METHODS AND LIMITATIONS", "STATUS:", "BASIS:", "CATEGORY:",
                    "VERIFICATION:", "MATH:", "APPROVED CORRECTIONS",
                    "KNOWN LIMITATIONS:", "SCOPE OF THIS EXPORT:", "PROVENANCE:"):
        assert heading in text, f"missing block: {heading}"


def test_the_standing_disclaimer_travels_in_every_channel():
    """Review **F-R3**: the "not a certification document" statement lived on the
    report's title page alone — the one page that does *not* travel with a
    forwarded CSV, deck or METHODS.txt. It is a required item of the statement
    (SUMMARY_REPORT.md §4.6 item 9), so it rides every channel the statement is
    wrapped for, and it leads the block rather than closing it: a reader who skims
    only the head of a stamped file must still meet it."""
    project = _project(_GA)
    text = methods_statement(project)
    assert STANDING_DISCLAIMER in text
    # Ahead of the load basis -- i.e. read before any number it qualifies.
    assert text.index(STANDING_DISCLAIMER) < text.index("BASIS:")
    # The statement writes one block per line, so each wrapper only prefixes it --
    # the sentence stays intact and greppable in the forwarded file itself.
    for wrapped in (csv_comment_block(project), bdf_comment_block(project)):
        assert STANDING_DISCLAIMER in wrapped


def test_the_title_page_and_the_statement_use_one_disclaimer_wording():
    """The cover quotes the constant instead of restating it: two wordings of the
    same disclaimer is two disclaimers, and a reader who spots the difference
    cannot tell which is current (the rule §4.6 applies to in-band caveats)."""
    from sloads.report.latex import render_report

    tex = render_report(_project(_GA))
    assert "Status." in tex
    # Escaped for LaTeX, so compare on the sentence that survives escaping intact.
    assert "not a certification document" in tex
    assert "See the methods and limitations section" in tex


def test_statement_states_ultimate_and_the_default_factor():
    text = methods_statement(_project(_GA))
    assert "ULTIMATE" in text
    assert "1.5" in text and "23.303" in text


#: The register of record. The guard below reads it rather than trusting the
#: tuple, so this path is part of the contract.
_REGISTER = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "docs", "20_theory", "02_approved_corrections.md",
)


def _register_headings(state):
    """``### `` headings under one ``## `` section of the register, suffix stripped.

    ``state`` is the section heading to read ("Register", "Considered and
    declined"). Parsing stops at the next ``## `` so that a *withdrawn* or
    *declined* entry can never be mistaken for an approved one -- which is the
    whole reason the register keeps them in separate sections.
    """
    out, inside = [], False
    with open(_REGISTER, encoding="utf-8") as fh:
        for line in fh:
            if line.startswith("## "):
                inside = line[3:].strip() == state
                continue
            if inside and line.startswith("### "):
                title = line[4:].strip()
                # Drop the trailing "*(approved 2026-08-17, issue #26)*" suffix:
                # the approval date is register metadata, not part of the entry's
                # identity, and pinning it here would make every re-approval a
                # code edit.
                out.append(title.split(" *(")[0].strip())
    return out


def test_statement_lists_every_approved_correction():
    """A deviation from the source manual that is not declared is invisible to the
    analyst — which defeats the point of approving it in the open.

    Checked against **the register**, not against the tuple the statement is
    rendered from. The previous form of this test looped over
    ``APPROVED_CORRECTIONS`` and asserted each key appeared in the text built out
    of it — circular, and blind to the four entries that were missing from the
    tuple for up to three weeks (2026-09-04 review R-3, issue #174). The register
    is the authority ``CLAUDE.md`` names, so the register is what the guard reads.
    """
    declared = [heading for heading, _, _ in APPROVED_CORRECTIONS]
    approved = _register_headings("Register")

    assert approved, "no ### entries parsed from the register's Register section"
    missing = [h for h in approved if h not in declared]
    assert not missing, (
        "approved correction(s) in the register but NOT declared in "
        "report/methods.APPROVED_CORRECTIONS, so they are invisible to every "
        f"analyst reading a stamped CSV, deck or report: {missing}"
    )
    extra = [h for h in declared if h not in approved]
    assert not extra, (
        "APPROVED_CORRECTIONS declares entries the register does not approve "
        f"under '## Register': {extra}"
    )
    assert declared == approved, (
        "the statement lists the approved corrections in a different order from "
        "the register; keep both in the register's own (chronological) order"
    )

    # ...and every declared entry actually reaches the deliverable.
    text = methods_statement(_project(_GA))
    for _, label, body in APPROVED_CORRECTIONS:
        assert f"  {label}: " in text, label
        assert body in text, label


def test_the_statement_declares_no_correction_the_register_declined():
    """The inverse drift, which the equality above cannot catch on its own.

    The register keeps *Withdrawn from scope* and *Considered and declined* as
    separate sections precisely because those are things sloads does **not** do.
    A heading that migrates from one of them into the statement would advertise a
    deviation the owner refused — worse than an undeclared one, because it is a
    false claim rather than a silent omission.
    """
    declared = {heading for heading, _, _ in APPROVED_CORRECTIONS}
    for state in ("Withdrawn from scope", "Considered and declined"):
        for heading in _register_headings(state):
            assert heading not in declared, (heading, state)


# --------------------------------------------------------------------------- #
# Limitations completeness (review F-R4)
# --------------------------------------------------------------------------- #
#: Every standing limitation the deliverable claims, by key. **This is the
#: completeness contract.** F-R4's finding was not a wrong sentence but a missing
#: one: the list was described as "every open caveat" and four were absent (the
#: fin-only lateral aero, the lumped aileron couple, the centreline clamp and the
#: flight-only body deck). Opening or closing a caveat now edits this set in the
#: same commit, and an omission is a red test rather than a silent claim of
#: completeness.
STANDING_LIMITATION_KEYS = {
    "control-surface-distributions",
    "export-case-filter",
    "flight-only-body-deck",
    "pressurization",
    "lateral-aero",
    "aileron-couple",
    "centerline-clamp",
    "engine-failure-propeller-only",
    # Decision G-9, opened with the ground families: ground and flight are
    # separate governing families and no envelope over both is claimed. Declared
    # here rather than left to be inferred from the absence of a comparison --
    # which is exactly the "silent claim of completeness" this set exists to end.
    "ground-flight-separate-families",
}


def test_the_standing_limitations_are_exactly_the_declared_set():
    keys = [key for key, _ in _standing_limitations()]
    assert len(keys) == len(set(keys)), keys
    assert set(keys) == STANDING_LIMITATION_KEYS


def test_every_standing_limitation_reaches_the_statement():
    """The list is only a claim until it travels: the block an analyst reads is
    what must carry each one, in every channel it is wrapped for."""
    text = methods_statement(_project(_GA))
    limitations = text[text.index("KNOWN LIMITATIONS:"):]
    for key, item in _standing_limitations():
        assert item in limitations, key


def test_pressurization_is_stated_as_an_exclusion_not_a_gap():
    """Decision **D-24** (2026-08-14): pressurization is permanently out of scope,
    so its limitation is a *scope statement*, not a pending-work note. The two read
    identically to a tool that only checks the key is present, and completely
    differently to the analyst deciding whether to wait for the next release — the
    old wording ("No pressurization load cases.") invited exactly that wait. Four
    shipped fixtures are pressurized airplanes, so the sentence has to carry its own
    finality wherever it lands."""
    text = dict(_standing_limitations())["pressurization"]
    assert "OUT OF SCOPE" in text
    assert "permanent exclusion" in text
    assert "23.365" in text
    # It travels, like every other standing limitation.
    assert text in methods_statement(_project(_GA))


def test_the_in_band_caveats_and_the_report_use_one_wording():
    """A caveat that reads one way on the deck or the case and another in the
    controlling document is two caveats, and a reader who spots the difference
    cannot tell which is current. Each of these is owned by the module that
    applies it and quoted, not paraphrased, by the report."""
    from sloads.export.sbeam_bridge import CENTERLINE_CLAMP_NOTE
    from sloads.modules.balance import AILERON_COUPLE_NOTE, LATERAL_AERO_NOTE
    from sloads.modules.one_engine_out import PROPELLER_ONLY_NOTE

    text = methods_statement(_project(_GA))
    for owner_note in (LATERAL_AERO_NOTE, AILERON_COUPLE_NOTE, CENTERLINE_CLAMP_NOTE,
                       PROPELLER_ONLY_NOTE):
        # From the second character: the notes are written to sit mid-sentence
        # in band, and one of them opens a report bullet, so only the case of the
        # first letter may differ. Every other character must match.
        assert owner_note[1:] in text, owner_note[:60]


def test_the_lateral_caveat_states_a_direction_per_degree_of_freedom():
    """The sentence shipped 2026-08-09 said `n_y` and the yaw acceleration were
    both OVER-stated and the inertia they drive therefore conservative. The yaw
    half is right; the `n_y` half is backwards -- at +beta the missing body and
    wing side force acts the same way as the fin's restoring load, so it ADDS and
    `|n_y|` is UNDER-stated, which makes the lateral translational inertia
    unconservative. Pinned per degree of freedom because the two directions
    differ and one sentence covering both is what got it wrong: a future edit
    that collapses them again fails here."""
    from sloads.modules.balance import LATERAL_AERO_NOTE

    note = LATERAL_AERO_NOTE
    assert "yaw acceleration is OVER-STATED" in note
    assert "n_y is UNDER-STATED" in note
    # ...and the consequence for the inertia, stated both ways round.
    assert "the inertia it drives is conservative" in note
    assert "NOT conservative" in note
    # L-7 (2026-08-17) closed the "unknown amount": the term is now computed and
    # the standing sentence defers the magnitude to the case, which states it
    # (decision L-7.16) -- so no fixed number, and no "unknown", lives here.
    assert "unknown amount" not in note
    assert "OFF by default" in note
    assert "estimated amount the case states" in note
    # It still travels, wording unchanged, into the controlling document.
    assert note[1:] in methods_statement(_project(_GA))


def test_the_statement_says_which_state_the_l7_term_is_in():
    """The controlling document names the term's state for the project it
    describes -- disabled on the shipped GA, enabled when the input says so."""
    from dataclasses import replace

    from sloads.models import LateralBodyAeroInput

    project = _project(_GA)
    assert "Lateral body aero (L-7) is DISABLED for this project" in methods_statement(project)
    project.aero_coeffs = replace(project.aero_coeffs,
                                  lateral_body_aero=LateralBodyAeroInput(enabled=True))
    text = methods_statement(project)
    assert "Lateral body aero (L-7) is ENABLED for this project (DATCOM" in text


def test_every_lateral_case_carries_one_of_the_two_l7_sentences():
    """Decision L-7.16: two stamped wordings, one per state of the term, both
    pinned so the deck can never claim the term was applied when it was not
    (or the reverse). The GA fixture ships with the term off, so its lateral
    cases carry the DISABLED sentence with the *estimated* effect; enabling it
    switches every one of them to the APPLIED sentence with the numbers."""
    from dataclasses import replace

    from sloads.models import LateralBodyAeroInput
    from sloads.modules.balance import build_balanced_cases, is_lateral

    project = _project(_GA)
    off = [c for c in build_balanced_cases(project) if is_lateral(c)]
    assert off
    for c in off:
        assert any("lateral body aero (L-7) DISABLED -- estimated for this case" in n
                   for n in c.notes), c.label
        assert not any("APPLIED" in n for n in c.notes), c.label
    project.aero_coeffs = replace(project.aero_coeffs,
                                  lateral_body_aero=LateralBodyAeroInput(enabled=True))
    on = [c for c in build_balanced_cases(project) if is_lateral(c)]
    for c in on:
        if c.beta_deg == 0.0:
            assert any("ENABLED but beta = 0" in n for n in c.notes), c.label
        else:
            assert any("lateral body aero (L-7) APPLIED" in n for n in c.notes), c.label
        assert not any("DISABLED" in n for n in c.notes), c.label


def test_the_assumed_tail_planform_reaches_the_statement():
    """Review **F-R4**: ``resolve_tail_planform`` marks a derived rectangle
    ASSUMED, and that marker reached the page, the CSV and the result and stopped
    — so the controlling document described the distribution as if the planform
    had been entered. No shipped fixture enters one, so this fires on the GA."""
    # ``ga6_normal`` entered its printed Appendix A empennage on 2026-08-30, so
    # the fixture no longer derives one. The marker is still the thing under
    # test, so the planforms are removed to produce a project that does.
    project = _project(_GA)
    project.geometry.surfaces[:] = [
        s for s in project.geometry.surfaces if s.name not in ("htail", "vtail")]
    text = methods_statement(project)
    assert "Horizontal tail planform ASSUMED" in text
    assert "Vertical tail planform ASSUMED" in text
    assert "not the surface's own" in text


def test_an_entered_tail_planform_states_no_assumption():
    """The other half of the contract: the caveat is conditional, so a project
    that enters the planform must not carry it. Guards against a standing
    sentence dressed up as a conditional one."""
    from sloads.tail_geometry import resolve_tail_planform

    project = _project(_GA)
    surfaces = project.geometry.surfaces if project.geometry is not None else []
    wing = next((s for s in surfaces if s.name == project.wing_mass.surface), None)
    assert wing is not None, "fixture has no wing surface to copy a planform from"
    entered = _htail_surface_from(wing, project)
    # One surface entered, the other derived -- which is what makes this the
    # *conditional* half of the contract. ``ga6_normal`` now enters both, so the
    # fin is removed to restore the mixed state the caveat has to distinguish.
    project.geometry.surfaces[:] = [
        s for s in project.geometry.surfaces if s.name not in (entered.name, "vtail")]
    project.geometry.surfaces.append(entered)
    assert resolve_tail_planform(project, "htail").assumed is False
    text = methods_statement(project)
    assert "Horizontal tail planform ASSUMED" not in text
    assert "Vertical tail planform ASSUMED" in text     # the fin is still derived


def _htail_surface_from(wing, project):
    """An 'htail' surface entry whose area and span match the entered scalars.

    ``resolve_tail_planform`` validates an entered polyline against the
    oracle-authoritative area/span (1 %), so the fixture cannot simply borrow the
    wing's geometry -- it has to be the rectangle the derivation would produce,
    entered explicitly.
    """
    import copy

    ht = project.tail_loads
    span_in, area_sqft = ht.htail_semispan_in, ht.htail_area_sqft
    chord = area_sqft * 144.0 / (2.0 * span_in)
    surface = copy.deepcopy(wing)
    surface.name = "htail"
    x_le = ht.xt25 - 0.25 * chord          # on the scalar 25 %-MAC station (Pri 1)
    surface.leading_edge = [(x_le, 0.0), (x_le, span_in)]
    surface.trailing_edge = [(x_le + chord, 0.0), (x_le + chord, span_in)]
    return surface


def test_verification_states_the_twin_cases_are_not_oracle_locked():
    """The oracle-status distinction is the single most load-bearing caveat in the
    document: Appendix B is not bundled, so twin results are closure-locked."""
    text = methods_statement(_project(_GA))
    assert "CLOSURE-LOCKED" in text and "NOT ORACLE-LOCKED" in text
    assert "Appendix B" in text


def test_concept_fixture_gets_the_caveat_and_the_ga_one_does_not():
    ga = methods_statement(_project(_GA))
    concept = methods_statement(_project(_CONCEPT))
    assert "UNVERIFIED EXTRAPOLATION" in concept
    assert "UNVERIFIED EXTRAPOLATION" not in ga
    assert "FAR 23 category" in ga


def test_deselected_cases_are_named_never_silently_dropped():
    text = methods_statement(_project(_GA), scope="governing case set",
                             deselected_case_ids=["W-03", "HT-02"])
    assert "DESELECTED" in text
    assert "W-03" in text and "HT-02" in text


def test_statement_is_deterministic():
    """Two renders must be byte-identical, or the diff between two report
    revisions is unreadable and the tests turn flaky. Nothing reads the clock."""
    p = _project(_GA)
    assert methods_statement(p, tool_version="0.3.0") == methods_statement(p, tool_version="0.3.0")
    stamped = methods_statement(p, generated="2026-08-04T00:00:00Z")
    assert "2026-08-04T00:00:00Z" in stamped
    assert "Generated:" not in methods_statement(p), "no timestamp unless the caller supplies one"


# --------------------------------------------------------------------------- #
# Comment-block wrappers
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize("wrapper,marker", [(csv_comment_block, "#"), (bdf_comment_block, "$")])
def test_every_line_of_a_comment_block_is_marked(wrapper, marker):
    """One unmarked line turns a comment block into a parse error."""
    block = wrapper(_project(_GA))
    lines = [ln for ln in block.split("\n") if ln]
    assert lines
    assert all(ln.startswith(marker) for ln in lines), [ln for ln in lines
                                                        if not ln.startswith(marker)]


def test_strip_comment_lines_is_the_inverse_for_readers():
    block = csv_comment_block(_project(_GA))
    assert strip_comment_lines(block + "a,b\n1,2\n") == "a,b\n1,2\n"


# --------------------------------------------------------------------------- #
# The channels
# --------------------------------------------------------------------------- #
def test_load_cases_csv_carries_the_stamp_and_still_parses():
    project = _project(_GA)
    results = run_all_modules(project)
    module = next(r for r in results if r.conditions)
    plain = io.load_cases_csv(module)
    stamped = io.load_cases_csv(module, header_comment=csv_comment_block(project))

    assert stamped.startswith("#")
    assert "ULTIMATE" in stamped
    # The payload is untouched: same rows, same header.
    assert strip_comment_lines(stamped) == plain
    rows_plain = list(_csv.DictReader(_io.StringIO(plain)))
    rows_stamped = list(_csv.DictReader(_io.StringIO(strip_comment_lines(stamped))))
    assert rows_stamped == rows_plain
    assert rows_plain, "fixture produced no rows to compare"


def test_span_load_csv_carries_the_stamp_and_still_parses():
    project, wing = _wing_net(_GA)
    plain = sb.span_load_csv(wing)
    stamped = sb.span_load_csv(wing, header_comment=csv_comment_block(project))
    assert stamped.startswith("#") and "ULTIMATE" in stamped
    # The file carries comment lines of its own (the moment-convention block,
    # note 46 OR-69), so the invariant is that the stamp disturbs nothing --
    # both sides are stripped, not just the stamped one.
    assert strip_comment_lines(stamped) == strip_comment_lines(plain)


def test_case_index_csv_carries_the_stamp():
    project = _project(_GA)
    if project.envelope is None:
        project.envelope = build_envelope(project)
    plain = sb.case_index_csv(project)
    stamped = sb.case_index_csv(project, header_comment=csv_comment_block(project))
    assert stamped.startswith("#")
    assert strip_comment_lines(stamped) == plain


def test_pandas_reads_a_stamped_csv_with_comment_marker():
    """The documented consumer contract: ``comment='#'``."""
    pd = pytest.importorskip("pandas")
    project = _project(_GA)
    results = run_all_modules(project)
    module = next(r for r in results if r.conditions)
    stamped = io.load_cases_csv(module, header_comment=csv_comment_block(project))
    df = pd.read_csv(_io.StringIO(stamped), comment="#")
    plain_df = pd.read_csv(_io.StringIO(io.load_cases_csv(module)))
    assert list(df.columns) == list(plain_df.columns)
    assert len(df) == len(plain_df)


def test_workbook_reader_skips_the_stamp():
    """``export/workbook._csv_to_df`` is an in-repo reader; G8.3 audited it."""
    pytest.importorskip("pandas")
    from sloads.export.workbook import _csv_to_df

    project = _project(_GA)
    results = run_all_modules(project)
    module = next(r for r in results if r.conditions)
    stamped = io.load_cases_csv(module, header_comment=csv_comment_block(project))
    df = _csv_to_df(stamped)
    plain = _csv_to_df(io.load_cases_csv(module))
    assert df is not None and plain is not None
    assert list(df.columns) == list(plain.columns)
    assert len(df) == len(plain)


def test_workbook_gains_a_methods_sheet():
    pytest.importorskip("openpyxl")
    pytest.importorskip("pandas")
    from openpyxl import load_workbook

    from sloads.export.workbook import build_workbook

    project = _project(_GA)
    data = build_workbook(
        {"Project": project.name}, {}, {}, "", {},
        methods=methods_statement(project),
    )
    wb = load_workbook(_io.BytesIO(data))
    assert "Methods" in wb.sheetnames
    text = "\n".join(str(c.value) for row in wb["Methods"].iter_rows() for c in row)
    assert "ULTIMATE" in text


def test_summary_report_carries_the_same_statement(tmp_path):
    """Step G8: the report is a channel like the others. Its §5 is the shared
    statement verbatim, so the document and the CSV/BDF files stamped beside it in
    the bundle cannot state different bases (SUMMARY_REPORT.md §4.6)."""
    from sloads.report.content import build_report
    from sloads.report.latex import render_report

    project = _project(_GA)
    doc = build_report(project)
    assert doc.methods == methods_statement(project)
    assert "ULTIMATE" in doc.section("Methods and limitations").body[0]
    # And it survives the trip through the renderer (escaped, not dropped).
    assert "ULTIMATE" in render_report(project)


if __name__ == "__main__":
    raise SystemExit(pytest.main([__file__, "-q"]))
