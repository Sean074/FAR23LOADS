"""Tests for the multi-sheet .xlsx workbook export (Step D8.2).

Builds the workbook from the same GA6 example the other export-bridge tests
use, re-opens it with openpyxl, and spot-checks sheet names/cell values against
the CSV strings that fed them -- guarding against mis-mapping, not a new oracle
(the underlying load values are already oracle-checked elsewhere).
"""

import io as _io
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl  # noqa: E402

from sloads import io, registry  # noqa: E402
from sloads.units import Channel, UnitSystem, deliverable_units, units_statement  # noqa: E402
from sloads.export import sbeam_bridge as sb  # noqa: E402
from sloads.export.workbook import build_workbook  # noqa: E402
from sloads.modules.aileron import build_aileron  # noqa: E402
from sloads.modules.body_loads import build_body_loads  # noqa: E402
from sloads.modules.flap import build_flap  # noqa: E402
from sloads.modules.net_loads import build_net_loads  # noqa: E402
from sloads.modules.tab import build_tabs  # noqa: E402
from sloads.modules.taildist import build_tail_chordwise  # noqa: E402

_EXAMPLES = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "examples")
_GA = os.path.join(_EXAMPLES, "ga6_normal.project.json")


def _try(fn, *args):
    """Mirror the live Export page's own ``_try`` (app/views/export_report.py):
    tolerate a module that has nothing to compute (an example missing an optional
    upstream slice) rather than requiring every fixture to be complete. GA6 now
    carries fuselage_mass (M2R-3), so body_loads computes; this stays defensive
    for the leaner examples that still don't."""
    try:
        return fn(*args)
    except (ValueError, ZeroDivisionError, KeyError, IndexError):
        return []


def _build(system=UnitSystem.IMPERIAL):
    project = io.load_project(_GA)
    module_results = registry.run_all_modules(project)
    module_csvs = {mr.module: io.load_cases_csv(mr, system=system) for mr in module_results}
    module_labels = {mr.module: mr.module for mr in module_results}

    net = build_net_loads(project)
    body = _try(build_body_loads, project)
    tail = _try(build_tail_chordwise, project)
    control = build_aileron(project) + build_flap(project) + build_tabs(project)

    case_index_csv = sb.case_index_csv_from(
        net.wing_net, body, tail, control,
        *(mr.conditions for mr in module_results),
    )
    span_csvs = {"Wing Span Loads": sb.span_load_csv(net.wing_net, system=system),
                 "Control Surface Loads": sb.control_surface_csv(control, system=system)}
    if body:
        span_csvs["Fuselage Span Loads"] = sb.body_span_load_csv(body, system=system)
    if tail:
        span_csvs["Tail Chordwise"] = sb.tail_chordwise_csv(tail, system=system)
    project_info = {"Name": project.name, "Engineer": project.engineer or "", "Date": project.date or ""}

    xlsx_bytes = build_workbook(project_info, module_csvs, module_labels, case_index_csv,
                                span_csvs, system=system)
    return project_info, module_csvs, case_index_csv, span_csvs, xlsx_bytes


def _sheet_rows(wb, name):
    """A data sheet's rows below its A1 unit statement -- header row first."""
    return list(wb[name].iter_rows(values_only=True))[1:]


def test_workbook_has_expected_sheets():
    _, module_csvs, case_index_csv, span_csvs, xlsx_bytes = _build()
    wb = openpyxl.load_workbook(_io.BytesIO(xlsx_bytes), read_only=True)
    names = set(wb.sheetnames)
    assert "Project" in names
    assert "Case Index" in names
    for module, csv_text in module_csvs.items():
        # A module whose standard cases-CSV is empty (e.g. body_loads, whose
        # output is the fuselage span distribution) gets no module sheet -- match
        # build_workbook's behaviour, as test_workbook_module_sheet_matches_csv_row_count does.
        if csv_text.strip():
            assert module in names
    for title in span_csvs:
        assert title in names


def test_workbook_project_sheet_values():
    project_info, *_, xlsx_bytes = _build()
    wb = openpyxl.load_workbook(_io.BytesIO(xlsx_bytes), read_only=True)
    ws = wb["Project"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("Field", "Value")
    values = dict(rows[1:])
    for field, value in project_info.items():
        # openpyxl reads an empty-string cell back as None.
        assert (values[field] or "") == value


def test_workbook_module_sheet_matches_csv_row_count():
    _, module_csvs, *_ , xlsx_bytes = _build()
    wb = openpyxl.load_workbook(_io.BytesIO(xlsx_bytes), read_only=True)
    for module, csv_text in module_csvs.items():
        if not csv_text.strip():
            continue
        n_csv_rows = len(csv_text.strip().splitlines()) - 1  # minus header
        n_sheet_rows = wb[module].max_row - 2  # minus the A1 unit statement and the header
        assert n_sheet_rows == n_csv_rows, module


def test_every_data_sheet_states_its_own_units_in_band():
    """Review m14: one workbook carries both channels of the selected system, so
    the unit statement is per sheet. The module sheets are the HUMAN channel and
    the span-load sheets are the SOLVER channel -- in SI those sets differ
    (`N·m`/`kPa` vs `N·mm`/`MPa`), and a sheet claiming the other one is a
    SUMMARY_REPORT.md 3.5/4.7 conformance failure, not a footnote."""
    _, module_csvs, _, span_csvs, xlsx_bytes = _build(UnitSystem.SI)
    wb = openpyxl.load_workbook(_io.BytesIO(xlsx_bytes), read_only=True)
    human = units_statement(deliverable_units(UnitSystem.SI, Channel.HUMAN))
    solver = units_statement(deliverable_units(UnitSystem.SI, Channel.SOLVER))
    assert human != solver, "the SI channels must differ or this test proves nothing"

    for module, csv_text in module_csvs.items():
        if not csv_text.strip():
            continue
        a1 = wb[module]["A1"].value
        assert human in a1 and solver not in a1, module
    for title in span_csvs:
        a1 = wb[title]["A1"].value
        assert solver in a1, title
    # The case index carries no load quantity, so it claims neither set.
    index_a1 = wb["Case Index"]["A1"].value
    assert "no load quantities" in index_a1
    assert human not in index_a1 and solver not in index_a1
    # ...and the Project sheet names both channels, not one.
    project_values = "\n".join(
        str(v) for row in wb["Project"].iter_rows(values_only=True) for v in row if v
    )
    assert human in project_values and solver in project_values


def test_workbook_header_row_sits_below_the_unit_statement():
    """The statement is an added row 1; row 2 is still the CSV's header row, so a
    reader (or `pd.read_excel(..., skiprows=1)`) gets the same table as before."""
    _, module_csvs, case_index_csv, span_csvs, xlsx_bytes = _build()
    wb = openpyxl.load_workbook(_io.BytesIO(xlsx_bytes), read_only=True)
    for title, csv_text in list(span_csvs.items()) + [("Case Index", case_index_csv)]:
        header = [c for c in _sheet_rows(wb, title)[0] if c is not None]
        expected = [h for h in csv_text.strip().splitlines()
                    if not h.startswith("#")][0].split(",")
        assert header == expected, title


def test_workbook_excludes_bdf_text():
    *_, xlsx_bytes = _build()
    wb = openpyxl.load_workbook(_io.BytesIO(xlsx_bytes), read_only=True)
    for name in wb.sheetnames:
        for row in wb[name].iter_rows(values_only=True):
            for cell in row:
                assert not (isinstance(cell, str) and cell.startswith("$ SLOADS")), name


if __name__ == "__main__":
    import traceback

    tests = [v for k, v in sorted(globals().items()) if k.startswith("test_") and callable(v)]
    failed = 0
    for t in tests:
        try:
            t()
            print(f"PASS {t.__name__}")
        except Exception:
            failed += 1
            print(f"FAIL {t.__name__}")
            traceback.print_exc()
    print(f"\n{len(tests) - failed}/{len(tests)} passed")
    sys.exit(1 if failed else 0)
